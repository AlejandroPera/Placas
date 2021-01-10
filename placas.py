rom pyppeteer import launch
import asyncio, shutil, os, schedule, time, openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl import load_workbook
from datetime import datetime
import pandas as pd
from subprocess import check_output, STDOUT

prime_file=r'S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\PlacasPrime.xlsx'
sky_path_xlsx=r'S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\Unidades SKY-DHL Actualizadas.xlsx'
new_sky_path=r'S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\placasSky.xlsx'
sky_download_path=r"C:\Users\aperalda\Downloads\Unidades SKY-DHL Actualizadas.xlsx"
lineas_RC_path=r'S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\LineasRC.xlsx'
prime_csv_path=r'S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\PlacasPrime.csv'

dups_ship=[]
total_shipments=0
num_placas=0
downloaded=0
both_lines_compared=[]

async def actualizacion_placas():
    try:
        browser = await launch(headless = False)
        page = await browser.newPage()
        page.setDefaultNavigationTimeout(60000);
        await page.goto('https://docs.google.com/LONKTODOWNLOAD', timeout=45000, waitUntil='networkidle2')
        await page.waitForSelector('#docs-file-menu')
        await page.click('#docs-file-menu')
        await page.waitFor(2000)
        await page.keyboard.press('ArrowDown')
        await page.waitFor(2000)
        await page.keyboard.press('Enter')
        await page.waitFor(2000)
        await page.keyboard.press('Enter')
        try:
            await page.waitFor(13000)
        except:
            pass
        await browser.close()
        global downloaded
        downloaded= 1
    except:
        print('Reintentado descarga...')
        try:
            await browser.close()
            return 0
        except:
            return 0

           

def move_file():
    shutil.move(sky_download_path, sky_path_xlsx)
   
    try:
        os.rename(sky_path_xlsx,new_sky_path)
    except FileExistsError:
        os.replace(sky_path_xlsx,new_sky_path)

def remove_duplicateds(duplicated_info):
    single_info=list(dict.fromkeys(duplicated_info))
    return single_info

 
def launchQlik(route, name, retries):
   now = datetime.now()
   cmd = r'"C:\Program Files\QlikView\Qv.exe" /r ' + route
   for i in range(retries):
        try:
            output = check_output(cmd, stderr=STDOUT, timeout = 5000)
            print('Generado',now.strftime("%Y-%m-%d %H:%M"),': ', name)
            return 1
        except:
            print('Timeout',now.strftime("%Y-%m-%d %H:%M"),': ', name)
   return 0  


def load_sky_info():
   
    wb = load_workbook(new_sky_path)       #Determina el numero de filas
    ws = wb.worksheets[0]
    pre_linea=list(ws.columns)[2]
    linea =[cell.value for cell in pre_linea if cell.value != None]
    lineas=remove_duplicateds(linea)
    pre_placas = list(ws.columns)[6]
    placas=[cell.value for cell in pre_placas if cell.value != None]
    sky_info=[lineas,placas]
    return sky_info

def load_RC_lineas():
    wb = load_workbook(lineas_RC_path)       #Determina el numero de filas
    ws = wb.worksheets[0]
    pre_linea=list(ws.columns)[1]
    linea =[cell.value for cell in pre_linea]
    return linea

def compare_Sky(sky_info, shipment_xid, lineas_RC, placas, wb,linea):
    wb.create_sheet('Sky')
    ws=wb.worksheets[1]
    ws.cell(1,1).value='Shipment'
    ws.cell(1,2).value='Placas'
    init_cell=2
    total_lineas=0

    lineas_sky=sky_info[0]
    placas_sky=sky_info[1]

    global both_lines_compared
    dups=0
    for value in range(len(placas)):
        if placas[value][0] in placas_sky:
            ws.cell(init_cell,2).value=placas[value][0]
            ws.cell(init_cell,1).value=shipment_xid[placas[value][1]]
            init_cell+=1
            if linea[placas[value][1]] not in lineas_RC:
                both_lines_compared.append([shipment_xid[placas[value][1]],linea[placas[value][1]], 'Sky'])
            elif linea[placas[value][1]] in lineas_RC:
                both_lines_compared.append([shipment_xid[placas[value][1]],linea[placas[value][1]], 'Ambas'])
   
    total_placas=len(ws['B'])-1
   

    ws.cell(1,9).value='% Placas'
    ws.cell(1,8).value='% Placas Real'
    ws.cell(1,7).value='% Restante'

    plates_percentage=(total_placas/num_placas)*100
    ws.cell(2,9).value=(total_placas/total_shipments)*100
    ws.cell(2,8).value=plates_percentage
    ws.cell(2,7).value=100-plates_percentage
    return 0

def compare_RC(RC_lineas, shipment_xid, lineas, wb):
    wb.create_sheet('RC')
    ws=wb.worksheets[2]
    ws.cell(1,1).value='Shipment'
    ws.cell(1,2).value='Linea'
    init_cell=2
    for value in range(len(lineas)):
        if lineas[value] in RC_lineas:
            ws.cell(init_cell,2).value=lineas[value]
            ws.cell(init_cell,1).value=shipment_xid[value]
            init_cell+=1

    total_lineas=len(ws['B'])-1
    lines_percentage=(total_lineas/total_shipments)*100
    ws.cell(1,9).value='% Lineas'
    ws.cell(1,8).value='% Restante'
    ws.cell(2,9).value=lines_percentage
    ws.cell(2,8).value=100-lines_percentage
    return 0


def compare_both(sky_lineas,RC_lineas, shipment_xid,lineas, wb):
    wb.create_sheet('Sky_RC')
    ws=wb.worksheets[3]
    ws.cell(1,1).value='Shipment'
    ws.cell(1,2).value='Linea'
    ws.cell(1,3).value='Plataforma'
    init_cell=2

    ships_gid=[ship[0] for ship in both_lines_compared]
 
    duplicateds=0
    global dups_ship
    for value in range(len(lineas)):
        if lineas[value] in RC_lineas and shipment_xid[value] not in ships_gid:
            ws.cell(init_cell,2).value=lineas[value]
            ws.cell(init_cell,1).value=shipment_xid[value]
            ws.cell(init_cell,3).value='RC'
            init_cell+=1
   
    for value in range(len(both_lines_compared)):
        ws.cell(init_cell,2).value=both_lines_compared[value][1]
        ws.cell(init_cell,1).value=both_lines_compared[value][0]
        ws.cell(init_cell,3).value=both_lines_compared[value][2]
        init_cell+=1
       
    total_lineas=len(ws['B'])-1
    lines_percentage=(total_lineas/total_shipments)*100
    ws.cell(1,9).value='% Viajes'
    ws.cell(1,8).value='% Restante'
    ws.cell(2,9).value=lines_percentage
    ws.cell(2,8).value=100-lines_percentage

    return 0

def graphs(wb):
    ws_sky=wb.worksheets[1]
    ws_RC=wb.worksheets[2]
    ws_both=wb.worksheets[3]

    chart_sky = PieChart()
    labels_Sky = Reference(ws_sky, min_col = 7, max_col=8, min_row = 1)
                     
    data_Sky = Reference(ws_sky, min_col = 7,max_col=8, min_row = 2)
    chart_sky.add_data(data_Sky, titles_from_data = True)
 
    # set labels in the chart object
    chart_sky.set_categories(labels_Sky)
   
    # set the title of the chart
    chart_sky.title = " Porcentaje de Placas "
    ws_sky.add_chart(chart_sky, "F9")

    chart_RC = PieChart()
    labels_RC = Reference(ws_RC, min_col = 8, max_col=9, min_row = 1)
                     
    data_RC = Reference(ws_RC, min_col = 8,max_col=9, min_row = 2)
    chart_RC.add_data(data_RC, titles_from_data = True)
 
    # set labels in the chart object
    chart_RC.set_categories(labels_RC)
   
    # set the title of the chart
    chart_RC.title = " Porcentaje de Lineas "
    ws_RC.add_chart(chart_RC, "I9")

    chart_both = PieChart()
    labels_both = Reference(ws_both, min_col = 8, max_col=9, min_row = 1)
                     
    data_both = Reference(ws_both, min_col = 8,max_col=9, min_row = 2)
    chart_both.add_data(data_both, titles_from_data = True)
 
    # set labels in the chart object
    chart_both.set_categories(labels_both)
   
    # set the title of the chart
    chart_both.title = " Porcentaje de Viajes "
    ws_both.add_chart(chart_both, "F9")

def compare_to_prime():
    sky=load_sky_info()
    RC=load_RC_lineas()
    read_file = pd.read_csv (prime_csv_path)
    read_file.to_excel (prime_file, index = None, header=True)
    wb = load_workbook(prime_file)       #Determina el numero de filas
    ws = wb.worksheets[0]

    global total_shipments
    total_shipments = ws.max_row
   
    pre_shipment_xid=list(ws.columns)[0]
    pre_linea=list(ws.columns)[3]
    pre_placas=list(ws.columns)[11]
    shipment_xid=[cell.value for cell in pre_shipment_xid]
    lineas=[cell.value for cell in pre_linea]
    placas=[[cell.value.replace(' ','').replace('-',''),index] for index, cell in enumerate(pre_placas) if cell.value != None]
    global num_placas
    num_placas =len(placas)
           
    compare_Sky(sky, shipment_xid, RC, placas, wb,lineas)
    compare_RC(RC, shipment_xid, lineas, wb)
    compare_both(sky[0],RC, shipment_xid,lineas, wb)

    graphs(wb)
    if datetime.today().weekday() == 0:
        wb.save('S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\placasViernes.xlsx')
    else:
        wb.save('S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\placasDiario.xlsx')
    wb.close()

def main():
    while downloaded==0:
        print('Inicializando descarga')
        asyncio.get_event_loop().run_until_complete(actualizacion_placas())
    print('Descarga terminada')
    move_file()
    # launchQlik(r'S:\TRANSPORTE\LPC\TEMP\Alex\RC_SKY\Placas\PlatesGeneration.qvw', 'PlatesGeneration', 3)
    # print('Generando reporte')
    # compare_to_prime()
    # print('Ejecución terminada')
#main()
schedule.every().day.at("07:30").do(main)

while True:
    schedule.run_pending()
    time.sleep(1)